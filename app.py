from flask import Flask, render_template, request
import openpyxl
from collections import defaultdict

app = Flask(__name__)

# Load the Excel file once when the app starts
workbook = openpyxl.load_workbook('https://github.com/beamingnp/website/blob/79dccfa57dd68e8b3cb04f18e3236b97c545f75b/pregnancy1.xlsx')
sheet = workbook.active
data = defaultdict(str)

# Preprocess the Excel data for faster searching
for row in range(2, sheet.max_row + 1):
    code = sheet.cell(row=row, column=1).value
    explanation = sheet.cell(row=row, column=2).value
    if code is not None and explanation is not None:
        if code not in data:
            data[code] = ""
        if data[code]:
            data[code] += "\n\n"
        data[code] += explanation

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/download_pdf')
def download_pdf():
    pdf_url = 'https://github.com/beamingnp/website/blob/79dccfa57dd68e8b3cb04f18e3236b97c545f75b/book-3-pregnancy.pdf'
    return render_template('index.html', pdf_url=pdf_url)

@app.route('/search_code', methods=['POST'])
def search_code():
    search_term = request.form['search_term'].upper()
    if search_term in data:
        results = [{"code": search_term, "explanation": data[search_term]}]
    else:
        results = []
    return render_template('index.html', results=results)

if __name__ == '__main__':
    app.run(debug=True)
